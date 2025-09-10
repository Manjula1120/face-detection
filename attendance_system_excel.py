import cv2
import datetime
import os
import argparse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

def load_labels(path="labels.txt"):
    label_map = {}
    with open(path, "r") as f:
        for line in f:
            if ":" in line:
                k, v = line.strip().split(":")
                label_map[int(k)] = v
    return label_map

def get_today_excel():
    today = datetime.date.today().strftime("%Y-%m-%d")
    filename = f"attendance_{today}.xlsx"

    if os.path.exists(filename):
        wb = load_workbook(filename)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Attendance"
        headers = ["Name", "Date", "Time"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        wb.save(filename)

    return filename, wb, sheet

def main():
    parser = argparse.ArgumentParser(description="Real-time Face Recognition Attendance (Excel)")
    parser.add_argument("--camera", type=int, default=0, help="Webcam index")
    parser.add_argument("--threshold", type=float, default=70.0, help="LBPH confidence threshold")
    args = parser.parse_args()

    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.read("face_recognizer.yml")
    label_map = load_labels("labels.txt")
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")

    attendance_file, workbook, sheet = get_today_excel()
    marked = set(sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1))

    cap = cv2.VideoCapture(args.camera)
    if not cap.isOpened():
        raise RuntimeError("❌ Could not open webcam. Try --camera 1 or --camera 2")

    print("[INFO] Press 'q' to quit.")
    while True:
        ret, frame = cap.read()
        if not ret:
            break

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.1, 5, minSize=(80, 80))

        for (x, y, w, h) in faces:
            roi = cv2.resize(gray[y:y+h, x:x+w], (200, 200))
            id_, conf = recognizer.predict(roi)

            if conf < args.threshold and id_ in label_map:
                name = label_map[id_]
                if name not in marked:
                    now = datetime.datetime.now()
                    sheet.append([name, now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")])
                    workbook.save(attendance_file)
                    marked.add(name)
                text, color = f"{name} ({int(conf)})", (0, 255, 0)
            else:
                text, color = "Unknown", (0, 0, 255)

            cv2.rectangle(frame, (x, y, x+w, y+h), color, 2)
            cv2.putText(frame, text, (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, color, 2)

        cv2.imshow("Attendance System", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()
    print(f"[INFO] Attendance saved to {attendance_file}")

if __name__ == "__main__":
    main()
